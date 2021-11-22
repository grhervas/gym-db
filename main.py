from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.orm.exc import NoResultFound

from models import (Program, Block, Workout, Workout_set,
                    Exercise, Log_workout, Log_set)

from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from bs4 import BeautifulSoup
import re
# from datetime import datetime


def get_data_from_html(file):
    """
    Gets planning data from html file and returns dict with microcycle.

        Parameters:
            file (str or path): html file containing the info

        Returns:
            micro_dict (dict): dict containing same info organised
    """
    with open(file) as html_file:
        soup = BeautifulSoup(html_file, "html.parser")

        sessions_list = []
        sessions = soup.find_all("div", class_="dia")
        for session in sessions:
            session_dict = {}

            workout_desc = session.find("div", class_="titulo").text.lower().strip()
            session_dict["workout_desc"] = workout_desc

            day = int(session.find("div", id="dia").text)
            month = int(session.find("div", id="mes").text)
            year = int(session.find("div", id="anyo").text)
            date = pd.Timestamp(day=day, month=month, year=year).date()
            session_dict["date_workout"] = date

            exercises_list = []
            exercises = (
                session.find("div", class_="cuerpo-boxdia")
                .find_all("div", re.compile("ejercicio.*"))
            )
            for exercise in exercises:
                exercise_dict = {}
                for element in exercise.find_all("div"):
                    exercise_dict[element.attrs['class'][0]] = element.text.strip()
                exercises_list.append(exercise_dict)
            session_dict["exercises"] = exercises_list

            sessions_list.append(session_dict)

        # # We have changed this in order to make it easier to manually modify Micro name
        # block_name = soup.find(id="microciclo").text.strip()
        if isinstance(file, Path):
            block_name = file.stem.strip()
        else:
            block_name = Path(file).stem.strip()
        micro_dict = {block_name: sessions_list}

    return micro_dict


def curate_exercises_data(exercises: list, col_names: list):
    """
    Apply some column transformations to make exercise data from workout
    ingestible by app.

        Parameters:
            exercises (list): list of exercises dicts
            col_names (list): list of standardized names to apply to dataframe

        Returns:
            df_exercises (pandas.DataFrame): dataframe containing workout exercises
                                             transformed
    """

    df_exercises = pd.DataFrame(exercises)
    # # Data quality criteria for workout session table
    # Standard names (only if matching length, else keep originals)
    if len(df_exercises.columns) == len(col_names):
        df_exercises.columns = col_names
    # Lowercase exercise names ("Ejercicio")
    df_exercises.iloc[:, 0] = df_exercises.iloc[:, 0].str.lower()
    # Convert to numbers
    # "Series"
    df_exercises.iloc[:, 1] = df_exercises.iloc[:, 1].str.extract(r"(\d+)").astype(int).values
    # "Cargas"
    df_exercises.iloc[:, 2] = df_exercises.iloc[:, 2].str.extract(r"(\d+)").astype(float).values
    # "Kilos"
    df_exercises.iloc[:, 3] = df_exercises.iloc[:, 3].str.extract(r"(\d+)").astype(float).values
    # "Repeticiones"
    df_exercises.iloc[:, 4] = df_exercises.iloc[:, 4].str.extract(r"(\d+)").astype(int).values
    # "RPE"
    df_exercises.iloc[:, 5] = df_exercises.iloc[:, 5].str.extract(r"(\d+)").astype(int).values
    # "Descanso"
    df_exercises.iloc[:, 6] = df_exercises.iloc[:, 6].str.extract(r"(\d+)").astype(float).values
    # In "Cargas (%)", "Kilos" and "Descanso (min)" change 0 --> NULL
    df_exercises.iloc[:, 2] = df_exercises.iloc[:, 2].replace(0.0, np.nan)
    df_exercises.iloc[:, 3] = df_exercises.iloc[:, 3].replace(0, np.nan)
    df_exercises.iloc[:, 6] = df_exercises.iloc[:, 6].replace(0.0, np.nan)

    return df_exercises


def add_block(session, source_file=None, program: str = None):
    """
    Adds block of program from personal coach html file

        Parameters:
            session (SQLAlchemy.session object)
            source_file (str or path): the .html file that contains the info
            program (str): the gym program description name
    """

    COL_NAMES = ["Ejercicio", "Series", "Cargas (%)",
                 "Kilos", "Repeticiones", "RPE", "Descanso (min)"]

    # Check if program description already exists
    if program:
        program = (
            session.query(Program)
            .filter(Program.program_desc == program)
            .one_or_none()
        )
        program_id = program.program_id
    # If not (or not provided), create new generic program
    if program is None:
        program = Program()
        session.add(program)
        # If new, the program_id will be last element added
        program_id = session.query(Program).count()

    block_dict = get_data_from_html(source_file)

    for block_name, workouts_list in block_dict.items():
        # Check if block already exist (matching both name and program)
        block = (
            session.query(Block)
            .filter(Block.block_desc == block_name, Block.program_id == program_id)
            .one_or_none()
        )
        if block:
            block_id = block.block_id
        # If not, create it
        if block is None:
            block = Block(block_desc=block_name, program_id=program_id)
            session.add(block)
            # If new, the block_id will be last element added
            block_id = session.query(Block).count()

            # The loading of workouts and sets is only done if block doesn't exist
            # If block already exists we could branch an update function inside
            # (inside this add_block() or outside as a different call)
            for wod in workouts_list:
                workout = Workout(workout_desc=wod["workout_desc"],
                                  block_id=block_id,
                                  date_workout=wod["date_workout"])
                session.add(workout)
                workout_id = session.query(Workout).count()
                df_exercises = curate_exercises_data(wod["exercises"], COL_NAMES)

                # If exercise is not in Exercise lookup table, add it previously
                data_ex = df_exercises["Ejercicio"].unique()
                db_ex = [i[0] for i in session.query(Exercise.exercise_desc).all()]
                for new_exercise in list(set(data_ex) - set(db_ex)):
                    exercise = Exercise(exercise_desc=new_exercise)
                    session.add(exercise)

                # Then, insert row by row the results (exploding for as many series
                # per exercise there are)
                for index, row in df_exercises.iterrows():
                    for wod_set in range(row["Series"]):
                        set_id = wod_set + 1
                        workout_set = Workout_set(
                            workout_id=workout_id,
                            exercise_id=session.query(Exercise.exercise_id)
                                               .filter_by(exercise_desc=row["Ejercicio"])
                                               .scalar(),
                            set_id=set_id,
                            no_reps=row["Repeticiones"],
                            weight=row["Kilos"],
                            perc_rm=row["Cargas (%)"] if row["Cargas (%)"] != np.nan else None,
                            max_rpe=row["RPE"],
                            rest_min=row["Descanso (min)"] if row["Descanso (min)"] != np.nan else None
                        )
                        session.add(workout_set)

    session.commit()


def generate_program_excel(session, program: int or str,
                           output_dir="/mnt/c/Users/gonza/OneDrive/Gym/routines_log/"):
    """
    Generates Excel file (.xlsx) with Program planning. Each program block
    is a different sheet with all the corresponding workouts.

        Parameters:
            session (SQLAlchemy.session object)
            program (int or str): Program identifier integer or description
                                  from database
            output_dir (str): Directory to store generated file
    """
    # If program description provided, get id
    if isinstance(program, str):
        program_id = (
            session.query(Program.program_id)
            .filter_by(program_desc=program)
            .scalar()
        )
    # If numeric or otherwise
    else:
        program_id = program

    # To check if valid id
    try:
        program = session.query(Program).filter_by(program_id=program_id).one()
    except NoResultFound:
        raise KeyError(f"Program_id ({program_id}) doesn't exist!")

    program_name = program.program_desc if program.program_desc else f"Program_{program.program_id}"

    file = output_dir + program_name + ".xlsx"

    if Path(file).is_file():
        # raise FileExistsError(f"{file.name} already exists in {file.parent}!")
        book = load_workbook(file)
    else:
        book = None

    with pd.ExcelWriter(file, engine="openpyxl") as writer:
        program_blocks = session.query(Block).filter_by(program_id=program_id).all()
        for program_block in program_blocks:
            block_name = program_block.block_desc
            # Load existing excel file into current if exists...
            if book:
                writer.book = book
            if block_name not in writer.book.sheetnames:
                start_row = 0
                for workout in program_block.workouts:
                    df_workout_header = pd.DataFrame([workout.date_workout, workout.workout_desc,
                                                      None, None, None],
                                                     index=["Fecha", "Descripción", "Duración (min)",
                                                            "RPE general", "Comentario general"])
                    df_workout = pd.read_sql(
                        session.query(Workout_set.workout_set_id.label("ID"),
                                      Exercise.exercise_desc.label("Ejercicio"),
                                      Workout_set.set_id.label("Serie"),
                                      Workout_set.no_reps.label("Repeticiones"),
                                      Workout_set.weight.label("Peso (kg)"),
                                      Workout_set.perc_rm.label("% 1RM"),
                                      Workout_set.min_rpe.label("RPE mín."),
                                      Workout_set.max_rpe.label("RPE máx."),
                                      Workout_set.rest_min.label("Descanso (min)"))
                        .join(Exercise.workout_sets)
                        .filter(Workout_set.workout_id == workout.workout_id)
                        .order_by(Workout_set.workout_set_id)
                        .statement,
                        session.bind)

                    # Add log fields (the No. Sets, No. Reps, Weight and RPE should be replaced
                    # if needed)
                    df_workout[["¿Hecho?", "RPE", "Comentarios"]] = None

                    df_workout_header.to_excel(writer, sheet_name=block_name,
                                               startrow=start_row,
                                               index=True, header=False)
                    start_row += df_workout_header.shape[0]
                    df_workout.to_excel(writer, sheet_name=block_name,
                                        startrow=start_row,
                                        index=False)
                    start_row += (df_workout.shape[0] + 2)


def load_log_data(session, log_file):
    """
    Load log data from Excel file (containing whole program) into db.

        Parameters:
            session (SQLAlchemy.session object)
            log_file (str or path): the Excel file that contains the info
    """

    # 1st. Get log file name to assign to correct Program
    program_desc = Path(log_file).stem

    if not session.query(Program).filter(Program.program_desc == program_desc).one_or_none():
        raise KeyError(f"No record for {program_desc}!")

    # 2nd. Iterate through blocks
    for block_desc, df_block in pd.read_excel(log_file, sheet_name=None, header=None).items():

        # Now we separate between header and body info (to log_workout and log_set, respectively)
        idx = df_block.index[df_block.isna().all(axis=1)].tolist()
        # We add the -1 to then add 1 in the loop and avoid including empty rows
        idx_mod = [-1] + idx + [len(df_block)]

        # Iterate through workouts
        for i in range(len(idx_mod)-1):
            df_wod = df_block.iloc[idx_mod[i]+1:idx_mod[i+1]]

            # Separate header (log_workout info)...
            df_wod_header = (df_wod.loc[df_wod.iloc[:, 2:].isna().all(axis=1)]
                                   .dropna(axis=1, how="all")
                                   .set_index(0).squeeze())
            # Check if it is even possible to have a record (past date condition)
            # if df_wod_header["Fecha"] <= datetime.today():
            if df_wod_header[["RPE general"]].notnull().all():
                # Get workout_id from Block and Program names and the order of workout in excel file
                workout_id = (session.query(Workout)
                                     .join(Block).join(Program)
                                     .filter(Block.block_desc == block_desc,
                                             Program.program_desc == program_desc)
                                     .order_by(Workout.workout_id)
                                     .all()[i].workout_id)

                # If log exists for workout_id, update the info
                log_workout = (session.query(Log_workout)
                                      .filter(Log_workout.workout_id == workout_id)
                                      .one_or_none())

                if log_workout:
                    log_workout.date_workout_done = df_wod_header["Fecha"]
                    log_workout.duration_min = df_wod_header["Duración (min)"]
                    log_workout.intensity = df_wod_header["RPE general"]
                    log_workout.comment_workout = df_wod_header["Comentario general"]
                # If not, insert the info
                else:
                    log_workout = Log_workout(workout_id=workout_id,
                                              date_workout_done=df_wod_header["Fecha"],
                                              duration_min=df_wod_header["Duración (min)"],
                                              intensity=df_wod_header["RPE general"],
                                              comment_workout=df_wod_header["Comentario general"])
                    session.add(log_workout)
                    # session.commit()

                # ... from body (log_set info)
                df_wod_exer = df_wod.loc[df_wod.iloc[:, 2:].notna().any(axis=1)]
                df_wod_exer.columns = df_wod_exer.iloc[0]
                df_wod_exer = df_wod_exer.iloc[1:]
                df_exer_done = df_wod_exer.loc[df_wod_exer[["¿Hecho?", "RPE"]].notnull().any(axis=1)]

                # Iterate through sets
                for _, row in df_exer_done.iterrows():
                    # If log exists for wod_set_id, update info
                    log_set = (session.query(Log_set)
                                      .filter(Log_set.workout_set_id == row["ID"])
                                      .one_or_none())
                    if log_set:
                        log_set.log_workout_id = log_workout.log_workout_id
                        log_set.no_reps_done = row["Repeticiones"]
                        log_set.weight_done = row["Peso (kg)"]
                        log_set.rpe_done = row["RPE"]
                        log_set.comment_set = row["Comentarios"]

                    # If not, insert info
                    else:
                        log_set = Log_set(workout_set_id=row["ID"],
                                          log_workout_id=log_workout.log_workout_id,
                                          no_reps_done=row["Repeticiones"],
                                          weight_done=row["Peso (kg)"],
                                          rpe_done=row["RPE"],
                                          comment_set=row["Comentarios"])
                        session.add(log_set)

    session.commit()


def main():
    """Main entry point of the program"""

    MACRO_NAME = "Macro Pisano"
    LOGS_DIR = "/mnt/c/Users/gonza/OneDrive/Gym/routines_log/"

    # Connect to the database using SQLAlchemy
    # sqlite_filepath = Path("./../gym_database.db").resolve()
    engine = create_engine(f"sqlite:///data/db/gym_database.db")
    Session = sessionmaker(bind=engine)
    session = Session()

    # Add blocks (html files) from "data/" folder
    html_files = [x for x in Path("data/").glob("*.html") if x.is_file()]
    for file in html_files:
        add_block(session, file, MACRO_NAME)

    # Create excel for macrocycle recording
    generate_program_excel(session, MACRO_NAME, LOGS_DIR)

    # Load excel records into db
    load_log_data(session, LOGS_DIR + MACRO_NAME + ".xlsx")


if __name__ == "__main__":
    main()
